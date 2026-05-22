"""
feedback_uret.py
───────────────────
MongoDB'deki yorumları okur → cihaz tespiti + kategori → JSON + Excel çıktısı

Yapı:
  Cihaz → Sorun Kategorisi → Yorumlar

Kullanım:
  python feedback_uret.py
  python feedback_uret.py --source sikayetvar --limit 500
  python feedback_uret.py --cikti ./raporlar/

DÜZELTİLEN HATALAR:
  - kategori_tespit() içinde normalize_complaint() çağrısı tuple döndürebiliyordu;
    _safe_normalize() ile sarıldı.
  - repair_feedback_klasoru() içindeki yorum_obj işleme mantığı güçlendirildi;
    hem dict ({"yorum": ..., "skor": ...}) hem de düz string formatı destekleniyor.
  - excel_kaydet() içinde yorum_obj str/dict ayrımı zaten vardı, ama
    feedback_kaydet()'e geçirilen veri formatı netleştirildi.
"""

import argparse
import json
import csv
import re
import os
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Optional

# ── Bağımlılık kontrolleri ──────────────────────────────
try:
    from pymongo import MongoClient
except ImportError:
    raise SystemExit("❌  pymongo yüklü değil: pip install pymongo")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("❌  openpyxl yüklü değil: pip install openpyxl")

try:
    from complaint_normalizer import normalize_complaint
    from cihaz_tespit import toplu_tespit, cihaz_tespit
    from yorum_kumeleme_v2 import konu_disi_mi
except ImportError:
    raise SystemExit("❌  cihaz_tespit.py veya complaint_normalizer.py bulunamadı.")


# ─────────────────────────────────────────────────────────
# YARDIMCI: Güvenli normalize
# ─────────────────────────────────────────────────────────

def _safe_normalize(text: str) -> str:
    """
    normalize_complaint()'in str döndürdüğünü garanti eder.
    Eski hatalı versiyonda tuple dönebiliyordu.
    """
    result = normalize_complaint(text)
    if isinstance(result, tuple):
        result = result[0] if result else ""
    return result if isinstance(result, str) else str(result)


# ═══════════════════════════════════════════════════════════
# 1. KATEGORİ SİSTEMİ
# ═══════════════════════════════════════════════════════════

_KATEGORILER = [
    "Batarya ve Şarj Sorunu",
    "Ekran ve Görünüm Sorunu",
    "Kamera Sorunu",
    "Yazılım ve Güncelleme Sorunu",
    "Servis ve Garanti Sorunu",
    "Performans ve Hız Sorunu",
    "Isınma Sorunu",
    "Ses ve Ses Kaydı Sorunu",
    "Ağ ve Bağlantı Sorunu",
    "Depolama Sorunu",
    "iCloud ve Hesap Sorunu",
    "Su ve Nem Hasarı",
    "Fiyat ve Ücret Şikayeti",
    "Kargo ve Teslimat Sorunu",
    "Diğer",
]

_EMOJI_MAP = {
    "Batarya ve Şarj Sorunu":       "🔋",
    "Ekran ve Görünüm Sorunu":      "📺",
    "Kamera Sorunu":                "📷",
    "Yazılım ve Güncelleme Sorunu": "⚙️",
    "Servis ve Garanti Sorunu":     "🏥",
    "Performans ve Hız Sorunu":     "⚡",
    "Isınma Sorunu":                "🔥",
    "Ses ve Ses Kaydı Sorunu":      "🔊",
    "Ağ ve Bağlantı Sorunu":        "📡",
    "Depolama Sorunu":              "💾",
    "iCloud ve Hesap Sorunu":       "☁️",
    "Su ve Nem Hasarı":             "💧",
    "Fiyat ve Ücret Şikayeti":      "💰",
    "Kargo ve Teslimat Sorunu":     "📦",
    "Diğer":                        "📌",
}

_KATEGORI_KURALLARI: list[tuple[str, str]] = [
    (r"batarya|pil\b|sarj|sarz|charger|battery|pil bi[td]|hizli bit", "Batarya ve Şarj Sorunu"),
    (r"ekran|dokunmatik|display|lcd|oled|catlak|kirik cam|titriyor|piksel|goruntu bozuk|touch", "Ekran ve Görünüm Sorunu"),
    (r"kamera|fotograf|video|lens|odak|blur|bulanik|cekim|selfie|arka kamera|on kamera", "Kamera Sorunu"),
    (r"guncelleme|update|ios|yazilim|uygulama|app|cokuyor|crash|boot|sistem hatasi|format|yeniden kur", "Yazılım ve Güncelleme Sorunu"),
    (r"servis|garanti|teknik destek|yetkili|onarim|tamir|apple store|musteri hizmet", "Servis ve Garanti Sorunu"),
    (r"yavas|kasiyor|donuyor|takiliyor|performans|hiz|acilmiyor|gec ac|lag|fps", "Performans ve Hız Sorunu"),
    (r"isin|sicak|yaniyor|overhe|termal|ates gibi", "Isınma Sorunu"),
    (r"ses\b|hoparlor|mikrofon|kulaklik|zil|audio|speaker|sound|duyulmuyor|ses yok|gurultu", "Ses ve Ses Kaydı Sorunu"),
    (r"wifi|wi-fi|bluetooth|sinyal|internet\b|baglanti|network|5g|4g|lte|hotspot|nfc|gsm|hat\b", "Ağ ve Bağlantı Sorunu"),
    (r"depolama|storage|hafiza|doldu|yer yok|gb\b|yedekleme alani|dosya sil", "Depolama Sorunu"),
    (r"icloud|apple id|hesap|sifre|senkronizasyon|sync|backup|yedek|oturum", "iCloud ve Hesap Sorunu"),
    (r"suya dus|islandi|nem|su\s*hasar|su\s*gecirmez|yagmur|sivi", "Su ve Nem Hasarı"),
    (r"fiyat|pahali|ucret|para|indirim|iade|fatura|fahis|odeme", "Fiyat ve Ücret Şikayeti"),
    (r"kargo|teslimat|gelmedi|gecikti|yanlis urun|paket|kutu|kurye|siparis", "Kargo ve Teslimat Sorunu"),
]


def kategori_tespit(yorum: str) -> str:
    """Yorum metninden sorun kategorisi tespit et (regex tabanlı)."""
    # DÜZELTME: _safe_normalize kullan
    m = _safe_normalize(yorum)
    m = unicodedata.normalize("NFKD", m).encode("ascii", "ignore").decode("ascii").lower()
    for pattern, kategori in _KATEGORI_KURALLARI:
        if re.search(pattern, m):
            return kategori
    return "Diğer"


# ═══════════════════════════════════════════════════════════
# 2. VERİ OKUMA + İŞLEME
# ═══════════════════════════════════════════════════════════

def veri_isle(
    connection_string: str = "mongodb://localhost:27017/",
    db_name: str = "apple_feedback_db",
    collection: str = "comments",
    source_filter: Optional[str] = None,
    limit: int = 0,
) -> dict:
    print("\n" + "═" * 65)
    print("  FEEDBACK OLUŞTURUCU")
    print("═" * 65)

    client = MongoClient(connection_string)
    db = client[db_name]
    col = db[collection]

    query: dict = {}
    if source_filter:
        query["source"] = source_filter

    cursor = col.find(query)
    if limit:
        cursor = cursor.limit(limit)

    docs = list(cursor)
    client.close()

    print(f"\n📥 {len(docs)} yorum okundu (kaynak: {source_filter or 'tümü'})\n")

    cihazlar: dict = defaultdict(
        lambda: {"toplam": 0, "kategoriler": defaultdict(lambda: {"toplam": 0, "yorumlar": []})}
    )
    apple_disi = 0
    belirsiz   = 0
    konu_disi  = 0
    islenen    = 0

    cleaned_rows = []
    for doc in docs:
        yorum_metni = (doc.get("normalized_comment") or doc.get("comment", "")).strip()
        if not yorum_metni:
            continue
        processing_text = yorum_metni or _safe_normalize(doc.get("comment", ""))
        cleaned_rows.append((doc, yorum_metni, processing_text))

    batch_size = 100
    for start in range(0, len(cleaned_rows), batch_size):
        batch = cleaned_rows[start:start + batch_size]
        batch_texts = [row[2] for row in batch]
        batch_results = toplu_tespit(batch_texts, embedding_fallback=False)

        for (doc, yorum_metni, processing_text), sonuc in zip(batch, batch_results):
            if not sonuc.is_apple:
                apple_disi += 1
                continue

            if sonuc.method == "none" or sonuc.confidence < 0.40:
                belirsiz += 1
                continue

            if konu_disi_mi(processing_text):
                konu_disi += 1
                continue

            normalized    = sonuc.normalized
            sonuc_method  = sonuc.method

            kategori = doc.get("category") or kategori_tespit(processing_text)

            tarih = doc.get("scraped_at")
            if hasattr(tarih, "strftime"):
                tarih_str = tarih.strftime("%Y-%m-%d")
            elif isinstance(tarih, dict) and "$date" in tarih:
                tarih_str = str(tarih["$date"])[:10]
            else:
                tarih_str = ""

            yorum_obj = {
                "id":     str(doc.get("_id", "")),
                "yorum":  processing_text,
                "kaynak": doc.get("source", ""),
                "tarih":  tarih_str,
                "detect": sonuc_method,
            }

            cihazlar[normalized]["toplam"] += 1
            cihazlar[normalized]["kategoriler"][kategori]["toplam"] += 1
            cihazlar[normalized]["kategoriler"][kategori]["yorumlar"].append(yorum_obj)

            islenen += 1
            if islenen % 500 == 0:
                print(f"  ✓ {islenen}/{len(cleaned_rows)} işlendi...")

    result_cihazlar = {}
    for cihaz, data in sorted(cihazlar.items(), key=lambda x: x[1]["toplam"], reverse=True):
        kategoriler_sorted = {}
        for kat in sorted(data["kategoriler"], key=lambda k: data["kategoriler"][k]["toplam"], reverse=True):
            kategoriler_sorted[kat] = dict(data["kategoriler"][kat])
        result_cihazlar[cihaz] = {
            "toplam":     data["toplam"],
            "kategoriler": kategoriler_sorted,
        }

    print(f"\n  ✅ Apple yorumları : {islenen}")
    print(f"  🚫 Apple dışı      : {apple_disi}")
    print(f"  ❓ Belirsiz         : {belirsiz}")
    print(f"  🧾 Konu dışı       : {konu_disi}")
    print(f"  📱 Benzersiz cihaz : {len(result_cihazlar)}\n")

    return {
        "meta": {
            "olusturulma_tarihi": datetime.now().isoformat(),
            "kaynak_filtresi":    source_filter or "tümü",
            "toplam_yorum":       islenen,
            "apple_disi":         apple_disi,
            "belirsiz":           belirsiz,
            "benzersiz_cihaz":    len(result_cihazlar),
        },
        "cihazlar": result_cihazlar,
    }


def _family_tahmin(normalized: str) -> str:
    n = normalized.lower()
    if "iphone" in n:   return "iPhone"
    if "ipad" in n:     return "iPad"
    if "mac" in n:      return "Mac"
    if "watch" in n:    return "Apple Watch"
    if "airpods" in n:  return "AirPods"
    if "homepod" in n:  return "HomePod"
    return "Diğer"


# ═══════════════════════════════════════════════════════════
# 3. JSON ÇIKTI
# ═══════════════════════════════════════════════════════════

def json_kaydet(veri: dict, dosya_yolu: str):
    with open(dosya_yolu, "w", encoding="utf-8") as f:
        json.dump(veri, f, ensure_ascii=False, indent=2)
    print(f"  💾 JSON kaydedildi → {dosya_yolu}")


def feedback_kaydet(cihaz: str, kumeler: dict):
    """
    Kategorileme sonuçlarını her cihaz için ayrı JSON + CSV dosyası olarak kaydet.

    kumeler formatı:
      {
        "🔋 Batarya ve Şarj Sorunu": {
          "yorumlar": ["yorum1", "yorum2"],   # str listesi
          "sorun_sayısı": 2,
          ...
        }
      }

    DÜZELTME: Yorum değerleri artık hem str hem de dict olabilir;
    her iki format normalize edilerek kaydedilir.
    """
    os.makedirs("feedback", exist_ok=True)

    json_file = f"feedback/{cihaz}_feedback.json"
    csv_file  = f"feedback/{cihaz}_feedback.csv"

    json_data = {}
    for kategori, veriler in kumeler.items():
        raw_yorumlar = veriler.get("yorumlar", [])
        # Normalize: her eleman str'e indirilir
        yorumlar_str = [
            (y.get("yorum", "") if isinstance(y, dict) else str(y))
            for y in raw_yorumlar
        ]
        json_data[kategori] = {
            "yorumlar":   yorumlar_str,
            "sorun_sayısı": veriler.get("sorun_sayısı", len(yorumlar_str)),
        }

    with open(json_file, "w", encoding="utf-8") as f:
        json.dump(json_data, f, ensure_ascii=False, indent=2)

    csv_rows = []
    for kategori, veriler in json_data.items():
        for yorum in veriler["yorumlar"]:
            csv_rows.append({"Kategori": kategori, "Yorum": yorum})

    if csv_rows:
        with open(csv_file, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=["Kategori", "Yorum"])
            writer.writeheader()
            writer.writerows(csv_rows)

    print(f"  ✓ {cihaz} güncellendi")


# ─────────────────────────────────────────────────────────
# repair_feedback_klasoru yardımcıları
# ─────────────────────────────────────────────────────────

_NORMALIZE_TRANSLATION = str.maketrans(
    {"ı": "i", "İ": "i", "ğ": "g", "ş": "s", "ç": "c", "ö": "o", "ü": "u"}
)

_EXPLICIT_PATTERNS: list[tuple[str, str]] = [
    ("Apple Watch Ultra 2",   r"apple\s*watch\s*ultra\s*[12]"),
    ("Apple Watch Ultra",     r"apple\s*watch\s*ultra"),
    ("Apple Watch SE",        r"apple\s*watch\s*se"),
    ("Apple Watch Series 10", r"apple\s*watch\s*(series\s*)?10"),
    ("Apple Watch Series 9",  r"apple\s*watch\s*(series\s*)?9"),
    ("Apple Watch Series 8",  r"apple\s*watch\s*(series\s*)?8"),
    ("Apple Watch",           r"apple\s*watch"),
    ("Apple Watch Ultra 2",   r"iphone\s*watch\s*ultra\s*[12]"),
    ("Apple Watch Ultra",     r"iphone\s*watch\s*ultra"),
    ("Apple Watch SE",        r"iphone\s*watch\s*se"),
    ("Apple Watch Series 10", r"iphone\s*watch\s*(series\s*)?10"),
    ("Apple Watch Series 9",  r"iphone\s*watch\s*(series\s*)?9"),
    ("Apple Watch Series 8",  r"iphone\s*watch\s*(series\s*)?8"),
    ("Apple Watch",           r"iphone\s*watch"),
    ("AirPods Pro 3",  r"airpods?\s*pro\s*3(nd)?"),
    ("AirPods Pro 2",  r"airpods?\s*pro\s*[23]"),
    ("AirPods Pro",    r"airpods?\s*pro"),
    ("AirPods Max",    r"airpods?\s*max"),
    ("AirPods",        r"airpods?\s*[234]"),
    ("AirPods",        r"airpods?"),
    ("Apple TV 4K",    r"apple\s*tv\s*4k"),
    ("Apple TV",       r"apple\s*tv"),
    ("HomePod Mini",   r"homepod\s*mini"),
    ("HomePod",        r"homepod"),
    ("MacBook Pro",    r"macbook\s*pro"),
    ("MacBook Air",    r"macbook\s*air"),
    ("MacBook",        r"macbook"),
    ("iMac",           r"imac"),
    ("iPad Pro 12.9",  r"i\s*pad\s*pro\s*1[12]\s*[.']?\s*9"),
    ("iPad Pro",       r"i\s*pad\s*pro\s*m\d|i\s*pad\s*pro"),
    ("iPad Air",       r"i\s*pad\s*air\s*m\d|i\s*pad\s*air\s*[2-6]|i\s*pad\s*air"),
    ("iPad Mini",      r"i\s*pad\s*mini\s*[2-7]|i\s*pad\s*mini"),
    ("iPad 10",        r"i\s*pad\s*10"),
    ("iPad",           r"i\s*pad\s*[6-9]|i\s*pad"),
    ("iPhone 17 Pro Max", r"i\s*ph[oa]?n[ei]?\s*17\s*pro\s*max"),
    ("iPhone 17 Pro",  r"i\s*ph[oa]?n[ei]?\s*17\s*pro"),
    ("iPhone 17 Plus", r"i\s*ph[oa]?n[ei]?\s*17\s*plus"),
    ("iPhone 17",      r"i\s*ph[oa]?n[ei]?\s*17"),
    ("iPhone 16 Pro Max", r"i\s*ph[oa]?n[ei]?\s*16\s*pro\s*max"),
    ("iPhone 16 Pro",  r"i\s*ph[oa]?n[ei]?\s*16\s*pro"),
    ("iPhone 16 Plus", r"i\s*ph[oa]?n[ei]?\s*16\s*plus"),
    ("iPhone 16",      r"i\s*ph[oa]?n[ei]?\s*16"),
    ("iPhone 15 Pro Max", r"i\s*ph[oa]?n[ei]?\s*15\s*pro\s*max"),
    ("iPhone 15 Pro",  r"i\s*ph[oa]?n[ei]?\s*15\s*pro"),
    ("iPhone 15 Plus", r"i\s*ph[oa]?n[ei]?\s*15\s*plus"),
    ("iPhone 15",      r"i\s*ph[oa]?n[ei]?\s*15"),
    ("iPhone 14 Pro Max", r"i\s*ph[oa]?n[ei]?\s*14\s*pro\s*max"),
    ("iPhone 14 Pro",  r"i\s*ph[oa]?n[ei]?\s*14\s*pro"),
    ("iPhone 14 Plus", r"i\s*ph[oa]?n[ei]?\s*14\s*plus"),
    ("iPhone 14",      r"i\s*ph[oa]?n[ei]?\s*14"),
    ("iPhone 13 Pro Max", r"i\s*ph[oa]?n[ei]?\s*13\s*pro\s*max"),
    ("iPhone 13 Pro",  r"i\s*ph[oa]?n[ei]?\s*13\s*pro"),
    ("iPhone 13 Mini", r"i\s*ph[oa]?n[ei]?\s*13\s*mini"),
    ("iPhone 13",      r"i\s*ph[oa]?n[ei]?\s*13"),
    ("iPhone 12 Pro Max", r"i\s*ph[oa]?n[ei]?\s*12\s*pro\s*max"),
    ("iPhone 12 Pro",  r"i\s*ph[oa]?n[ei]?\s*12\s*pro"),
    ("iPhone 12 Mini", r"i\s*ph[oa]?n[ei]?\s*12\s*mini"),
    ("iPhone 12",      r"i\s*ph[oa]?n[ei]?\s*12"),
    ("iPhone 11 Pro Max", r"i\s*ph[oa]?n[ei]?\s*11\s*pro\s*max"),
    ("iPhone 11 Pro",  r"i\s*ph[oa]?n[ei]?\s*11\s*pro"),
    ("iPhone 11",      r"i\s*ph[oa]?n[ei]?\s*11"),
    ("iPhone XS Max",  r"i\s*ph[oa]?n[ei]?\s*x\s*s\s*max"),
    ("iPhone XS",      r"i\s*ph[oa]?n[ei]?\s*x\s*s"),
    ("iPhone XR",      r"i\s*ph[oa]?n[ei]?\s*x\s*r"),
    ("iPhone X",       r"i\s*ph[oa]?n[ei]?\s*x\b"),
    ("iPhone SE 3",    r"i\s*ph[oa]?n[ei]?\s*se\s*[34]"),
    ("iPhone SE",      r"i\s*ph[oa]?n[ei]?\s*se"),
    ("iPhone 8 Plus",  r"i\s*ph[oa]?n[ei]?\s*8\s*plus"),
    ("iPhone 8",       r"i\s*ph[oa]?n[ei]?\s*8"),
    ("iPhone 7 Plus",  r"i\s*ph[oa]?n[ei]?\s*7\s*plus"),
    ("iPhone 7",       r"i\s*ph[oa]?n[ei]?\s*7"),
    ("iPhone 6s Plus", r"i\s*ph[oa]?n[ei]?\s*6s\s*plus"),
    ("iPhone 6s",      r"i\s*ph[oa]?n[ei]?\s*6s"),
    ("iPhone 6 Plus",  r"i\s*ph[oa]?n[ei]?\s*6\s*plus"),
    ("iPhone 6",       r"i\s*ph[oa]?n[ei]?\s*6"),
    ("iPhone",         r"i\s*ph[oa]?n[ei]"),
]

_GENERIC_STEMS = {
    "iPhone", "iPad", "MacBook", "Mac",
    "Apple Watch", "AirPods", "Apple TV", "HomePod", "iMac",
}


def _normalize_text(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text)
    normalized = normalized.replace("İ", "i").replace("i̇", "i")
    normalized = normalized.translate(_NORMALIZE_TRANSLATION)
    return normalized.encode("ascii", "ignore").decode("ascii").lower()


def _explicit_family_matches(text: str) -> list[tuple[int, str]]:
    normalized = _normalize_text(text)
    matches: list[tuple[int, str]] = []
    for stem, pattern in _EXPLICIT_PATTERNS:
        for match in re.finditer(pattern, normalized):
            matches.append((match.start(), stem))
    matches.sort(key=lambda item: item[0])
    return matches


def _explicit_target(text: str, source_stem: str) -> str | None:
    explicit = _explicit_family_matches(text)
    if not explicit:
        return None
    specific_matches = [item for item in explicit if item[1] not in _GENERIC_STEMS]
    if not specific_matches:
        return None
    source_is_generic = source_stem in _GENERIC_STEMS
    if not source_is_generic:
        for _, stem in specific_matches:
            if stem == source_stem:
                return source_stem
    return specific_matches[0][1]


def repair_feedback_klasoru(feedback_dir: str = "feedback") -> None:
    """
    Mevcut feedback dosyalarını tekrar tarayıp açıkça görünen cihaz ailelerine
    göre yeniden yazar.

    DÜZELTME: yorum_obj hem dict hem de str olabilir; her iki format da işlenir.
    """
    klasor = Path(feedback_dir)
    json_dosyalar = sorted(p for p in klasor.glob("*_feedback.json") if p.is_file())

    yeniden_dagitilan: dict = defaultdict(
        lambda: defaultdict(lambda: {"yorumlar": [], "sorun_sayısı": 0})
    )
    tasinan  = defaultdict(int)
    korunan  = defaultdict(int)

    for dosya in json_dosyalar:
        source_stem = dosya.name[: -len("_feedback.json")]
        with dosya.open("r", encoding="utf-8") as handle:
            data = json.load(handle)

        if not isinstance(data, dict):
            continue

        for kategori, payload in data.items():
            if not isinstance(payload, dict):
                continue
            yorumlar = payload.get("yorumlar", [])
            if not isinstance(yorumlar, list):
                continue

            yorum_listesi = []
            yorum_objs = []
            explicit_targets = []

            for yorum_obj in yorumlar:
                if isinstance(yorum_obj, dict):
                    yorum = str(yorum_obj.get("yorum", "")).strip()
                else:
                    yorum = str(yorum_obj).strip()

                if not yorum:
                    continue

                yorum_listesi.append(yorum)
                yorum_objs.append(yorum)
                explicit_targets.append(_explicit_target(yorum, source_stem))

            batch_size = 100
            for start in range(0, len(yorum_listesi), batch_size):
                batch_yorumlar = yorum_listesi[start:start + batch_size]
                batch_resultler = toplu_tespit(batch_yorumlar, embedding_fallback=False)

                for j, (yorum, sonuc) in enumerate(zip(batch_yorumlar, batch_resultler)):
                    explicit_target = explicit_targets[start + j]
                    if explicit_target:
                        hedef = explicit_target
                    else:
                        if not sonuc.is_apple or sonuc.normalized == "apple_disi":
                            hedef = "bilinmeyen"
                        elif sonuc.method == "regex" and sonuc.confidence >= 0.85 and sonuc.normalized:
                            hedef = sonuc.normalized
                        else:
                            hedef = source_stem

                    yeniden_dagitilan[hedef][kategori]["yorumlar"].append(yorum)
                    yeniden_dagitilan[hedef][kategori]["sorun_sayısı"] += 1

                    if hedef == source_stem:
                        korunan[source_stem] += 1
                    else:
                        tasinan[hedef] += 1

    print(f"İşlenecek dosya sayısı: {len(json_dosyalar)}")
    print(f"Hedef dosya sayısı: {len(yeniden_dagitilan)}")

    for cihaz in sorted(yeniden_dagitilan):
        toplam = sum(kat["sorun_sayısı"] for kat in yeniden_dagitilan[cihaz].values())
        print(
            f"{cihaz}_feedback -> toplam={toplam}, "
            f"taşınan={tasinan.get(cihaz, 0)}, "
            f"korunan={korunan.get(cihaz, 0)}"
        )

    for cihaz, kumeler in yeniden_dagitilan.items():
        feedback_kaydet(cihaz, kumeler)

    print("Feedback dosyaları güncellendi.")


# ═══════════════════════════════════════════════════════════
# 4. EXCEL ÇIKTI
# ═══════════════════════════════════════════════════════════

_RENK = {
    "baslik_bg": "1C1C1E",
    "baslik_fg": "FFFFFF",
    "cihaz_bg":  "2C2C2E",
    "cihaz_fg":  "FFFFFF",
    "kat_bg":    "F2F2F7",
    "kat_fg":    "1C1C1E",
    "yorum_bg":  "FFFFFF",
    "yorum_fg":  "3A3A3C",
    "ozet_bg":   "007AFF",
    "ozet_fg":   "FFFFFF",
    "alt_bg":    "F9F9F9",
    "border":    "D1D1D6",
}

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def _font(bold=False, color="000000", size=10, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic, name="Arial")

def _border_bottom(color="D1D1D6") -> Border:
    s = Side(style="thin", color=color)
    return Border(bottom=s)

def _border_all(color="D1D1D6") -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def excel_kaydet(veri: dict, dosya_yolu: str):
    wb = openpyxl.Workbook()
    _ozet_sayfasi(wb, veri)
    _cihaz_sayfalari(wb, veri)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(dosya_yolu)
    print(f"  📊 Excel kaydedildi → {dosya_yolu}")


def _ozet_sayfasi(wb: openpyxl.Workbook, veri: dict):
    ws = wb.create_sheet("📊 Özet", 0)
    ws.sheet_view.showGridLines = False

    meta    = veri.get("meta", {}) or {}
    cihazlar = veri.get("cihazlar", {})

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    row = 1
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "Apple Feedback Raporu"
    ws[f"A{row}"].font      = _font(bold=True, color=_RENK["baslik_fg"], size=14)
    ws[f"A{row}"].fill      = _fill(_RENK["baslik_bg"])
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 32
    row += 1

    ws.merge_cells(f"A{row}:E{row}")
    olusturulma = meta.get("olusturulma_tarihi", datetime.now().isoformat())[:16]
    kaynak      = meta.get("kaynak_filtresi", "tümü")
    ws[f"A{row}"] = f"Oluşturulma: {olusturulma}  |  Kaynak: {kaynak}"
    ws[f"A{row}"].font      = _font(color="AAAAAA", size=9, italic=True)
    ws[f"A{row}"].fill      = _fill(_RENK["cihaz_bg"])
    ws[f"A{row}"].alignment = Alignment(horizontal="center")
    row += 2

    stats = [
        ("Toplam Apple Yorumu",  meta.get("toplam_yorum", 0)),
        ("Apple Dışı (taşındı)", meta.get("apple_disi", 0)),
        ("Belirsiz",             meta.get("belirsiz", 0)),
        ("Benzersiz Cihaz",      meta.get("benzersiz_cihaz", 0)),
    ]
    for label, val in stats:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = _font(bold=True, size=10)
        ws[f"B{row}"] = val
        ws[f"B{row}"].font = _font(size=10, color=_RENK["ozet_bg"])
        ws[f"B{row}"].alignment = Alignment(horizontal="center")
        row += 1

    row += 1
    basliklar = ["Cihaz", "Toplam Yorum", "Kategori Sayısı", "En Çok Sorun", "Adet"]
    for col_idx, h in enumerate(basliklar, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font      = _font(bold=True, color=_RENK["ozet_fg"], size=10)
        cell.fill      = _fill(_RENK["ozet_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _border_all()
    ws.row_dimensions[row].height = 20
    row += 1

    for i, (cihaz, data) in enumerate(cihazlar.items()):
        en_cok_kat  = max(data["kategoriler"], key=lambda k: data["kategoriler"][k]["toplam"], default="-")
        en_cok_adet = data["kategoriler"].get(en_cok_kat, {}).get("toplam", 0)
        bg = _RENK["alt_bg"] if i % 2 == 0 else _RENK["yorum_bg"]
        satir_data = [cihaz, data["toplam"], len(data["kategoriler"]), en_cok_kat, en_cok_adet]
        for col_idx, val in enumerate(satir_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font      = _font(size=9)
            cell.fill      = _fill(bg)
            cell.alignment = Alignment(
                horizontal="center" if col_idx > 1 else "left", vertical="center"
            )
            cell.border = _border_bottom()
        row += 1


def _cihaz_sayfalari(wb: openpyxl.Workbook, veri: dict):
    cihazlar = veri["cihazlar"]
    aileler: dict = defaultdict(dict)
    for cihaz, data in cihazlar.items():
        aileler[_family_tahmin(cihaz)][cihaz] = data

    for aile, aile_cihazlari in sorted(aileler.items()):
        sayfa_adi = aile[:28].replace("/", "-").replace(":", "")
        ws = wb.create_sheet(sayfa_adi)
        ws.sheet_view.showGridLines = False

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 80
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 12

        row = 1
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = f"{aile} — Yorum Detayları"
        ws[f"A{row}"].font      = _font(bold=True, color=_RENK["baslik_fg"], size=13)
        ws[f"A{row}"].fill      = _fill(_RENK["baslik_bg"])
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 28
        row += 2

        col_basliklar = ["Cihaz / Kategori", "Adet", "Yorum", "Kaynak", "Tarih"]
        for col_idx, h in enumerate(col_basliklar, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font      = _font(bold=True, color=_RENK["ozet_fg"], size=10)
            cell.fill      = _fill(_RENK["ozet_bg"])
            cell.alignment = Alignment(horizontal="center")
            cell.border    = _border_all()
        ws.row_dimensions[row].height = 18
        row += 1

        for cihaz, data in sorted(aile_cihazlari.items(), key=lambda x: x[1]["toplam"], reverse=True):
            ws.merge_cells(f"A{row}:E{row}")
            ws[f"A{row}"] = f"  📱 {cihaz}  —  {data['toplam']} yorum"
            ws[f"A{row}"].font      = _font(bold=True, color=_RENK["cihaz_fg"], size=11)
            ws[f"A{row}"].fill      = _fill(_RENK["cihaz_bg"])
            ws[f"A{row}"].alignment = Alignment(vertical="center")
            ws.row_dimensions[row].height = 22
            row += 1

            for kategori, kat_data in data["kategoriler"].items():
                emoji = _EMOJI_MAP.get(kategori, "📌")
                ws[f"A{row}"] = f"    {emoji} {kategori}"
                ws[f"A{row}"].font = _font(bold=True, size=10, color=_RENK["kat_fg"])
                ws[f"A{row}"].fill = _fill(_RENK["kat_bg"])
                ws[f"B{row}"] = kat_data["toplam"]
                ws[f"B{row}"].font      = _font(bold=True, size=10)
                ws[f"B{row}"].fill      = _fill(_RENK["kat_bg"])
                ws[f"B{row}"].alignment = Alignment(horizontal="center")
                for col_idx in range(3, 6):
                    ws.cell(row=row, column=col_idx).fill = _fill(_RENK["kat_bg"])
                ws.row_dimensions[row].height = 18
                row += 1

                for j, yorum_obj in enumerate(kat_data["yorumlar"]):
                    bg = _RENK["yorum_bg"] if j % 2 == 0 else _RENK["alt_bg"]

                    if isinstance(yorum_obj, dict):
                        yorum_metni = yorum_obj.get("yorum", "")
                        kaynak      = yorum_obj.get("kaynak", "")
                        tarih       = yorum_obj.get("tarih", "")
                    else:
                        yorum_metni = str(yorum_obj)
                        kaynak      = ""
                        tarih       = ""

                    ws[f"A{row}"] = ""
                    ws[f"B{row}"] = ""
                    ws[f"C{row}"] = yorum_metni
                    ws[f"D{row}"] = kaynak
                    ws[f"E{row}"] = tarih

                    for col_idx in range(1, 6):
                        cell = ws.cell(row=row, column=col_idx)
                        cell.fill      = _fill(bg)
                        cell.font      = _font(size=9, color=_RENK["yorum_fg"])
                        cell.alignment = Alignment(
                            wrap_text=(col_idx == 3), vertical="top"
                        )
                        cell.border = _border_bottom()
                    ws.row_dimensions[row].height = 40
                    row += 1

            row += 1


# ═══════════════════════════════════════════════════════════
# 5. ANA FONKSİYON
# ═══════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Apple feedback raporu oluştur")
    parser.add_argument("--connection",  default="mongodb://localhost:27017/")
    parser.add_argument("--db",          default="apple_feedback_db")
    parser.add_argument("--collection",  default="comments")
    parser.add_argument("--source",      default=None)
    parser.add_argument("--limit",       default=0, type=int)
    parser.add_argument("--cikti",       default=".")
    args = parser.parse_args()

    os.makedirs(args.cikti, exist_ok=True)

    veri = veri_isle(
        connection_string=args.connection,
        db_name=args.db,
        collection=args.collection,
        source_filter=args.source,
        limit=args.limit,
    )

    zaman     = datetime.now().strftime("%Y%m%d_%H%M")
    json_yol  = os.path.join(args.cikti, f"feedback_{zaman}.json")
    excel_yol = os.path.join(args.cikti, f"feedback_{zaman}.xlsx")

    print("📝 Dosyalar oluşturuluyor...\n")
    json_kaydet(veri, json_yol)
    excel_kaydet(veri, excel_yol)

    print("\n" + "═" * 65)
    print("  ✅ TAMAMLANDI")
    print(f"  JSON  : {json_yol}")
    print(f"  Excel : {excel_yol}")
    print("═" * 65 + "\n")


if __name__ == "__main__":
    main()